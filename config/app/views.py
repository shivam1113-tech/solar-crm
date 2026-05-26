from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from .models import Lead, Customer, Project, Quote, Invoice, FollowUp
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponseForbidden
import random
import csv


# ================= AUTH =================

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ================= DASHBOARD =================

@login_required
def dashboard(request):
    today = timezone.now().date()

    if request.user.is_superuser:
        all_leads = Lead.objects.all()
    else:
        all_leads = Lead.objects.filter(assigned_to=request.user)

    total_leads = all_leads.count()
    new_leads = all_leads.filter(status="New").count()
    contacted = all_leads.filter(status="Contacted").count()
    qualified = all_leads.filter(status="Qualified").count()
    proposal = all_leads.filter(status="Proposal").count()

    total_deals = Quote.objects.filter(status="Accepted").count()
    revenue = Invoice.objects.filter(status="Paid").values_list('amount', flat=True)
    total_revenue = sum(revenue)

    recent_leads = all_leads.order_by('-created_at')[:5]
    upcoming_followups = FollowUp.objects.filter(
        follow_up_date__gte=today, done=False
    ).order_by('follow_up_date', 'follow_up_time')[:5]

    return render(request, 'dashboard.html', {
        'total_leads': total_leads,
        'new_leads': new_leads,
        'total_deals': total_deals,
        'total_revenue': total_revenue,
        'leads_new': new_leads,
        'leads_contacted': contacted,
        'leads_qualified': qualified,
        'leads_proposal': proposal,
        'recent_leads': recent_leads,
        'upcoming_followups': upcoming_followups,
    })


# ================= LEADS =================

@login_required
def leads(request):
    if request.user.is_superuser:
        leads = Lead.objects.all().order_by('-created_at')
        employees = User.objects.filter(is_superuser=False, is_active=True)
    else:
        leads = Lead.objects.filter(assigned_to=request.user).order_by('-created_at')
        employees = []
    return render(request, 'leads.html', {'leads': leads, 'employees': employees})


@login_required
def view_lead(request, id):
    lead = get_object_or_404(Lead, id=id)
    if not request.user.is_superuser and lead.assigned_to != request.user:
        messages.error(request, "You don't have access to this lead")
        return redirect('leads')
    return render(request, 'view_lead.html', {'lead': lead})


@login_required
def add_lead(request):
    employees = User.objects.filter(is_superuser=False, is_active=True)
    if request.method == "POST":
        assigned_id = request.POST.get('assigned_to')
        # ── FIXED: clean solar fields ──
        monthly_bill = request.POST.get('monthly_bill') or None
        solar_kwh_required = request.POST.get('solar_kwh_required') or None

        Lead.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            description=request.POST.get('description'),
            status=request.POST.get('status'),
            budget=request.POST.get('budget') or 0,
            monthly_bill=monthly_bill,
            solar_kwh_required=solar_kwh_required,
            assigned_to=User.objects.filter(id=assigned_id).first() if assigned_id else None
        )
        messages.success(request, "Lead added successfully")
        return redirect('leads')
    return render(request, 'lead_form.html', {'title': 'Add Lead', 'employees': employees})


@login_required
def edit_lead(request, id):
    lead = get_object_or_404(Lead, id=id)
    if not request.user.is_superuser and lead.assigned_to != request.user:
        messages.error(request, "You don't have access to this lead")
        return redirect('leads')
    employees = User.objects.filter(is_superuser=False, is_active=True)
    if request.method == "POST":
        assigned_id = request.POST.get('assigned_to')
        lead.name = request.POST.get('name')
        lead.email = request.POST.get('email')
        lead.phone = request.POST.get('phone')
        lead.description = request.POST.get('description')
        lead.status = request.POST.get('status')
        lead.budget = request.POST.get('budget') or 0
        # ── FIXED: save solar fields ──
        lead.monthly_bill = request.POST.get('monthly_bill') or None
        lead.solar_kwh_required = request.POST.get('solar_kwh_required') or None
        lead.assigned_to = User.objects.filter(id=assigned_id).first() if assigned_id else None
        lead.save()
        messages.success(request, "Lead updated successfully")
        return redirect('view_lead', id=id)
    return render(request, 'lead_form.html', {'lead': lead, 'title': 'Edit Lead', 'employees': employees})


@login_required
def delete_lead(request, id):
    if not request.user.is_superuser:
        messages.error(request, "Only admin can delete leads")
        return redirect('leads')
    lead = get_object_or_404(Lead, id=id)
    lead.delete()
    messages.success(request, "Lead deleted successfully")
    return redirect('leads')


@login_required
def convert_to_customer(request, id):
    if not request.user.is_superuser:
        messages.error(request, "Only admin can convert leads to customers")
        return redirect('view_lead', id=id)
    lead = get_object_or_404(Lead, id=id)
    if Customer.objects.filter(email=lead.email).exists():
        messages.warning(request, f"{lead.name} is already a customer!")
        return redirect('view_lead', id=id)
    Customer.objects.create(lead=lead, name=lead.name, email=lead.email, phone=lead.phone)
    lead.status = "Won"
    lead.save()
    messages.success(request, f"{lead.name} converted to customer!")
    return redirect('customers')


# ================= CUSTOMERS =================

@login_required
def customers(request):
    if request.user.is_superuser:
        customers = Customer.objects.all().order_by('-created_at')
    else:
        customers = Customer.objects.filter(lead__assigned_to=request.user).order_by('-created_at')
    return render(request, 'customers.html', {'customers': customers})


@login_required
def view_customer(request, id):
    customer = get_object_or_404(Customer, id=id)
    return render(request, 'view_customer.html', {'customer': customer})


@login_required
def add_customer(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    if request.method == "POST":
        Customer.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            address=request.POST.get('address'),
        )
        messages.success(request, "Customer added successfully")
        return redirect('customers')
    return render(request, 'customer_form.html', {'title': 'Add Customer'})


@login_required
def edit_customer(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    customer = get_object_or_404(Customer, id=id)
    if request.method == "POST":
        customer.name = request.POST.get('name')
        customer.email = request.POST.get('email')
        customer.phone = request.POST.get('phone')
        customer.address = request.POST.get('address')
        customer.save()
        messages.success(request, "Customer updated successfully")
        return redirect('customers')
    return render(request, 'customer_form.html', {'customer': customer, 'title': 'Edit Customer'})


@login_required
def delete_customer(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    customer = get_object_or_404(Customer, id=id)
    customer.delete()
    messages.success(request, "Customer deleted")
    return redirect('customers')


# ================= PROJECTS =================

@login_required
def projects(request):
    if request.user.is_superuser:
        projects = Project.objects.all().order_by('-created_at')
    else:
        my_customers = Customer.objects.filter(lead__assigned_to=request.user)
        projects = Project.objects.filter(customer__in=my_customers).order_by('-created_at')
    return render(request, 'projects.html', {'projects': projects})


@login_required
def view_project(request, id):
    project = get_object_or_404(Project, id=id)
    return render(request, 'view_project.html', {'project': project})


@login_required
def add_project(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    customers = Customer.objects.all()
    if request.method == "POST":
        Project.objects.create(
            customer=get_object_or_404(Customer, id=request.POST.get('customer')),
            title=request.POST.get('title'),
            description=request.POST.get('description'),
            status=request.POST.get('status'),
            start_date=request.POST.get('start_date') or None,
            end_date=request.POST.get('end_date') or None,
        )
        messages.success(request, "Project added successfully")
        return redirect('projects')
    return render(request, 'project_form.html', {'title': 'Add Project', 'customers': customers})


@login_required
def edit_project(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    project = get_object_or_404(Project, id=id)
    customers = Customer.objects.all()
    if request.method == "POST":
        project.customer = get_object_or_404(Customer, id=request.POST.get('customer'))
        project.title = request.POST.get('title')
        project.description = request.POST.get('description')
        project.status = request.POST.get('status')
        project.start_date = request.POST.get('start_date') or None
        project.end_date = request.POST.get('end_date') or None
        project.save()
        messages.success(request, "Project updated")
        return redirect('projects')
    return render(request, 'project_form.html', {'project': project, 'title': 'Edit Project', 'customers': customers})


@login_required
def delete_project(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    project = get_object_or_404(Project, id=id)
    project.delete()
    messages.success(request, "Project deleted")
    return redirect('projects')


# ================= QUOTES =================

@login_required
def quotes(request):
    if request.user.is_superuser:
        quotes = Quote.objects.all().order_by('-created_at')
    else:
        quotes = Quote.objects.filter(lead__assigned_to=request.user).order_by('-created_at')
    return render(request, 'quotes.html', {'quotes': quotes})


@login_required
def view_quote(request, id):
    quote = get_object_or_404(Quote, id=id)
    return render(request, 'view_quote.html', {'quote': quote})


@login_required
def add_quote(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    leads = Lead.objects.all()
    if request.method == "POST":
        Quote.objects.create(
            lead=get_object_or_404(Lead, id=request.POST.get('lead')),
            amount=request.POST.get('amount') or 0,
            status=request.POST.get('status'),
            notes=request.POST.get('notes'),
        )
        messages.success(request, "Quote added successfully")
        return redirect('quotes')
    return render(request, 'quote_form.html', {'title': 'Add Quote', 'leads': leads})


@login_required
def edit_quote(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    quote = get_object_or_404(Quote, id=id)
    leads = Lead.objects.all()
    if request.method == "POST":
        quote.lead = get_object_or_404(Lead, id=request.POST.get('lead'))
        quote.amount = request.POST.get('amount') or 0
        quote.status = request.POST.get('status')
        quote.notes = request.POST.get('notes')
        quote.save()
        messages.success(request, "Quote updated")
        return redirect('quotes')
    return render(request, 'quote_form.html', {'quote': quote, 'title': 'Edit Quote', 'leads': leads})


@login_required
def delete_quote(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    quote = get_object_or_404(Quote, id=id)
    quote.delete()
    messages.success(request, "Quote deleted")
    return redirect('quotes')


# ================= INVOICES =================

@login_required
def invoices(request):
    if not request.user.is_superuser:
        messages.error(request, "Access denied")
        return redirect('dashboard')
    invoices = Invoice.objects.all().order_by('-created_at')
    return render(request, 'invoices.html', {'invoices': invoices})


@login_required
def view_invoice(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    invoice = get_object_or_404(Invoice, id=id)
    return render(request, 'view_invoice.html', {'invoice': invoice})


@login_required
def add_invoice(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    customers = Customer.objects.all()
    if request.method == "POST":
        Invoice.objects.create(
            customer=get_object_or_404(Customer, id=request.POST.get('customer')),
            amount=request.POST.get('amount') or 0,
            status=request.POST.get('status'),
            due_date=request.POST.get('due_date') or None,
        )
        messages.success(request, "Invoice added successfully")
        return redirect('invoices')
    return render(request, 'invoice_form.html', {'title': 'Add Invoice', 'customers': customers})


@login_required
def delete_invoice(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    invoice = get_object_or_404(Invoice, id=id)
    invoice.delete()
    messages.success(request, "Invoice deleted")
    return redirect('invoices')


# ================= FOLLOW UPS =================

@login_required
def followups(request):
    if request.user.is_superuser:
        followups = FollowUp.objects.all().order_by('follow_up_date', 'follow_up_time')
    else:
        followups = FollowUp.objects.filter(lead__assigned_to=request.user).order_by('follow_up_date', 'follow_up_time')
    return render(request, 'followups.html', {'followups': followups})


@login_required
def view_followup(request, id):
    followup = get_object_or_404(FollowUp, id=id)
    if not request.user.is_superuser and followup.lead.assigned_to != request.user:
        messages.error(request, "Access denied")
        return redirect('followups')
    return render(request, 'view_followup.html', {'followup': followup})


@login_required
def add_followup(request):
    if request.user.is_superuser:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(assigned_to=request.user)
    if request.method == "POST":
        lead_id = request.POST.get('lead')
        FollowUp.objects.create(
            lead=get_object_or_404(Lead, id=lead_id),
            title=request.POST.get('title'),
            follow_up_date=request.POST.get('follow_up_date'),
            follow_up_time=request.POST.get('follow_up_time') or None,
            notes=request.POST.get('notes'),
        )
        messages.success(request, "Follow up added successfully")
        return redirect('followups')
    return render(request, 'followup_form.html', {'title': 'Add Follow Up', 'leads': leads})


@login_required
def delete_followup(request, id):
    if not request.user.is_superuser:
        messages.error(request, "Only admin can delete follow ups")
        return redirect('followups')
    followup = get_object_or_404(FollowUp, id=id)
    followup.delete()
    messages.success(request, "Follow up deleted")
    return redirect('followups')


@login_required
def toggle_followup(request, id):
    followup = get_object_or_404(FollowUp, id=id)
    followup.done = not followup.done
    followup.save()
    from django.http import JsonResponse
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.META.get('HTTP_X_CSRFTOKEN'):
        return JsonResponse({'success': True, 'done': followup.done})
    return redirect('followups')


@login_required
def edit_followup(request, id):
    followup = get_object_or_404(FollowUp, id=id)
    if not request.user.is_superuser and followup.lead.assigned_to != request.user:
        messages.error(request, "Access denied")
        return redirect('followups')
    leads = Lead.objects.all() if request.user.is_superuser else Lead.objects.filter(assigned_to=request.user)
    if request.method == "POST":
        followup.title = request.POST.get('title')
        followup.lead = get_object_or_404(Lead, id=request.POST.get('lead'))
        followup.follow_up_date = request.POST.get('follow_up_date')
        followup.follow_up_time = request.POST.get('follow_up_time') or None
        followup.notes = request.POST.get('notes')
        followup.save()
        messages.success(request, "Follow up updated successfully")
        return redirect('followups')
    return render(request, 'followup_form.html', {'title': 'Edit Follow Up', 'followup': followup, 'leads': leads})


# ================= IMPORT =================

@login_required
def import_data(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admin can import data")
        return redirect('dashboard')
    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "No file uploaded")
            return redirect('import_data')
        filename = file.name.lower()
        count = 0
        try:
            if filename.endswith('.csv'):
                decoded = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded)
                for row in reader:
                    row = {k.strip().lower(): v.strip() for k, v in row.items()}
                    Lead.objects.create(
                        name=row.get('name', ''), email=row.get('email', ''),
                        phone=row.get('phone', ''), description=row.get('description', ''),
                        status=row.get('status', 'New'), budget=row.get('budget', 0) or 0,
                    )
                    count += 1
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(cell.value).strip().lower() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    Lead.objects.create(
                        name=str(data.get('name') or ''), email=str(data.get('email') or ''),
                        phone=str(data.get('phone') or ''), description=str(data.get('description') or ''),
                        status=str(data.get('status') or 'New'), budget=data.get('budget') or 0,
                    )
                    count += 1
            else:
                messages.error(request, "Unsupported file. Use CSV or Excel.")
                return redirect('import_data')
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
            return redirect('import_data')
        messages.success(request, f"Successfully imported {count} lead(s)!")
        return redirect('leads')
    return render(request, 'import.html')


# ================= EMPLOYEES =================

@login_required
def employees(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    employees = User.objects.filter(is_superuser=False, is_active=True).order_by('-date_joined')
    return render(request, 'employees.html', {'employees': employees})


@login_required
def add_employee(request):
    if not request.user.is_superuser:
        return redirect('dashboard')
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return render(request, 'add_employee.html', {'error': 'Username already exists'})
        User.objects.create_user(username=username, email=email, password=password, is_staff=False, is_superuser=False)
        messages.success(request, f"Employee '{username}' created successfully!")
        return redirect('employees')
    return render(request, 'add_employee.html')


@login_required
def delete_employee(request, id):
    if not request.user.is_superuser:
        return redirect('dashboard')
    user = get_object_or_404(User, id=id, is_superuser=False)
    user.delete()
    messages.success(request, "Employee deleted")
    return redirect('employees')


# ================= FORGOT PASSWORD =================

def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return render(request, 'forgot.html', {'message': 'No account found with this email'})
        otp = str(random.randint(100000, 999999))
        request.session['reset_otp'] = otp
        request.session['reset_email'] = email
        try:
            import ssl, smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            msg = MIMEMultipart()
            msg['From'] = settings.EMAIL_HOST_USER
            msg['To'] = email
            msg['Subject'] = 'Solar CRM - Password Reset OTP'
            body = f'Hi {user.username},\n\nYour OTP for password reset is: {otp}\n\nValid for 10 minutes.\n\nIf you did not request this, ignore this email.'
            msg.attach(MIMEText(body, 'plain'))
            context = ssl.create_default_context()
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE
            with smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT) as server:
                server.ehlo()
                server.starttls(context=context)
                server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
                server.sendmail(settings.EMAIL_HOST_USER, email, msg.as_string())
        except Exception as e:
            return render(request, 'forgot.html', {'message': f'Failed to send email: {str(e)}'})
        return redirect('verify_reset_otp')
    return render(request, 'forgot.html')


def verify_reset_otp(request):
    if not request.session.get('reset_otp'):
        return redirect('forgot_password')
    email = request.session.get('reset_email', '')
    if request.method == "POST":
        user_otp = request.POST.get('otp')
        if user_otp == request.session.get('reset_otp'):
            del request.session['reset_otp']
            return redirect('reset_password')
        return render(request, 'forgot_verify.html', {'message': 'Invalid OTP. Please try again.', 'email': email})
    return render(request, 'forgot_verify.html', {'email': email})


def reset_password(request):
    if not request.session.get('reset_email'):
        return redirect('forgot_password')
    if request.method == "POST":
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        if password1 != password2:
            return render(request, 'reset_password.html', {'message': 'Passwords do not match'})
        if len(password1) < 6:
            return render(request, 'reset_password.html', {'message': 'Password must be at least 6 characters'})
        email = request.session.get('reset_email')
        try:
            user = User.objects.get(email=email)
            user.set_password(password1)
            user.save()
            request.session.flush()
            messages.success(request, "Password reset successful! Please login.")
            return redirect('login')
        except User.DoesNotExist:
            return render(request, 'reset_password.html', {'message': 'User not found'})
    return render(request, 'reset_password.html')


# ================= AJAX VIEWS =================

from django.http import JsonResponse
from django.views.decorators.http import require_POST

@login_required
def ajax_dashboard_stats(request):
    if request.user.is_superuser:
        all_leads = Lead.objects.all()
    else:
        all_leads = Lead.objects.filter(assigned_to=request.user)
    data = {
        'total_leads': all_leads.count(),
        'new_leads': all_leads.filter(status='New').count(),
        'contacted': all_leads.filter(status='Contacted').count(),
        'qualified': all_leads.filter(status='Qualified').count(),
        'proposal': all_leads.filter(status='Proposal').count(),
        'total_deals': Quote.objects.filter(status='Accepted').count(),
        'total_revenue': sum(Invoice.objects.filter(status='Paid').values_list('amount', flat=True)),
    }
    return JsonResponse(data)


@login_required
@require_POST
def ajax_change_lead_status(request, id):
    if not request.user.is_superuser:
        lead = get_object_or_404(Lead, id=id, assigned_to=request.user)
    else:
        lead = get_object_or_404(Lead, id=id)
    new_status = request.POST.get('status')
    valid = ['New', 'Contacted', 'Qualified', 'Proposal', 'Won', 'Lost']
    if new_status not in valid:
        return JsonResponse({'success': False, 'error': 'Invalid status'})
    lead.status = new_status
    lead.save()
    return JsonResponse({'success': True, 'status': new_status})


@login_required
@require_POST
def ajax_delete_lead(request, id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    lead = get_object_or_404(Lead, id=id)
    lead.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def ajax_add_lead(request):
    # ── FIXED: now saves monthly_bill and solar_kwh_required too ──
    try:
        assigned_id = request.POST.get('assigned_to')
        monthly_bill = request.POST.get('monthly_bill') or None
        solar_kwh_required = request.POST.get('solar_kwh_required') or None

        lead = Lead.objects.create(
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'New'),
            budget=request.POST.get('budget') or 0,
            monthly_bill=monthly_bill,
            solar_kwh_required=solar_kwh_required,
            assigned_to=User.objects.filter(id=assigned_id).first() if assigned_id else None
        )
        return JsonResponse({'success': True, 'lead': {
            'id': lead.id, 'name': lead.name, 'email': lead.email,
            'phone': lead.phone, 'status': lead.status,
            'budget': str(lead.budget), 'date': lead.created_at.strftime('%d %b %Y'),
            'monthly_bill': str(lead.monthly_bill or ''),
            'solar_kwh_required': str(lead.solar_kwh_required or ''),
        }})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def ajax_notifications(request):
    today = timezone.now().date()
    if request.user.is_superuser:
        upcoming = FollowUp.objects.filter(follow_up_date=today, done=False).values('id', 'title', 'follow_up_time', 'lead__name')[:5]
        new_leads = Lead.objects.filter(status='New').order_by('-created_at').values('id', 'name', 'created_at')[:5]
    else:
        upcoming = FollowUp.objects.filter(follow_up_date=today, done=False, lead__assigned_to=request.user).values('id', 'title', 'follow_up_time', 'lead__name')[:5]
        new_leads = Lead.objects.filter(status='New', assigned_to=request.user).order_by('-created_at').values('id', 'name', 'created_at')[:5]
    return JsonResponse({
        'upcoming_count': upcoming.count(), 'upcoming': list(upcoming),
        'new_leads_count': new_leads.count(), 'new_leads': list(new_leads),
        'total_notifications': upcoming.count() + new_leads.count()
    })


@login_required
def ajax_live_search(request):
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'results': []})
    if request.user.is_superuser:
        leads = Lead.objects.filter(name__icontains=query) | Lead.objects.filter(email__icontains=query) | Lead.objects.filter(phone__icontains=query)
    else:
        leads = Lead.objects.filter(assigned_to=request.user).filter(name__icontains=query)
    results = []
    for lead in leads[:8]:
        results.append({'id': lead.id, 'name': lead.name, 'email': lead.email, 'status': lead.status, 'type': 'lead'})
    customers = Customer.objects.filter(name__icontains=query)[:4] if request.user.is_superuser else []
    for c in customers:
        results.append({'id': c.id, 'name': c.name, 'email': c.email, 'status': 'Customer', 'type': 'customer'})
    return JsonResponse({'results': results})